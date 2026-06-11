import os
import re
import tempfile
from fastapi import HTTPException, status
from typing import TYPE_CHECKING, Optional

if TYPE_CHECKING:
    from markitdown import MarkItDown
else:
    try:
        from markitdown import MarkItDown
    except ImportError:
        MarkItDown = None

class DocumentService:
    def __init__(self):
        self._parser: Optional[MarkItDown] = None
        if MarkItDown is not None:
            try:
                self._parser = MarkItDown()
            except Exception:
                pass

    def get_parser(self) -> MarkItDown:
        if self._parser is None:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Microsoft MarkItDown is not loaded or installed. Ensure markitdown package is properly configured."
            )
        return self._parser

    def process_excel(self, file_content: bytes, fast_mode: bool = True) -> tuple[str, list[str]]:
        import io
        import openpyxl
        warnings = []
        markdown_sheets = []
        
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True, read_only=True)
        except Exception as e:
            raise ValueError(f"Failed to open Excel workbook: {str(e)}")
            
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            last_non_empty_row_idx = -1
            collected_rows = []
            
            for row_idx, r in enumerate(sheet.iter_rows(values_only=True)):
                is_empty = True
                if r:
                    for cell_val in r:
                        if cell_val is not None and str(cell_val).strip() != "":
                            is_empty = False
                            break
                        
                if not is_empty:
                    last_non_empty_row_idx = row_idx
                    
                if not fast_mode or row_idx < 5000:
                    collected_rows.append(r)
                    
            if last_non_empty_row_idx == -1:
                continue
                
            total_non_empty_rows = last_non_empty_row_idx + 1
            
            if fast_mode and total_non_empty_rows > 5000:
                collected_rows = collected_rows[:5000]
                warn_msg = f"Sheet {sheet_name}: 5000 of {total_non_empty_rows} rows processed (Fast Mode)"
                warnings.append(warn_msg)
            else:
                collected_rows = collected_rows[:total_non_empty_rows]
                if fast_mode:
                    warnings.append(f"Sheet {sheet_name}: {total_non_empty_rows} rows processed (Fast Mode)")
                else:
                    warnings.append(f"Sheet {sheet_name}: {total_non_empty_rows} rows processed")
                    
            max_cols = max(len(r) for r in collected_rows) if collected_rows else 0
            if max_cols == 0:
                continue
                
            sheet_md = [f"### Sheet: {sheet_name}"]
            for idx, r in enumerate(collected_rows):
                cells = list(r) + [None] * (max_cols - len(r))
                formatted_cells = []
                for cell_val in cells:
                    if cell_val is None:
                        formatted_cells.append("")
                    else:
                        formatted_cells.append(str(cell_val).replace("\n", " ").replace("\r", " ").strip())
                        
                sheet_md.append("| " + " | ".join(formatted_cells) + " |")
                if idx == 0:
                    sheet_md.append("| " + " | ".join(["---"] * max_cols) + " |")
                    
            markdown_sheets.append("\n".join(sheet_md))
            
        combined_md = "\n\n".join(markdown_sheets)
        return combined_md, warnings

    def convert_to_markdown(self, file_content: bytes, extension: str, fast_mode: bool = True) -> tuple[str, list[str]]:
        warnings = []
        if extension.lower() == ".xlsx":
            try:
                combined_md, excel_warnings = self.process_excel(file_content, fast_mode)
                return combined_md, excel_warnings
            except Exception:
                pass

        parser = self.get_parser()
        
        temp_fd, temp_path = tempfile.mkstemp(suffix=extension)
        try:
            with os.fdopen(temp_fd, 'wb') as tmp:
                tmp.write(file_content)
            result = parser.convert(temp_path)
            return result.text_content or "", warnings
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Failed to parse document: {str(e)}"
            )
        finally:
            if os.path.exists(temp_path):
                try:
                    os.unlink(temp_path)
                except OSError:
                    pass

    def markdown_to_plain_text(self, markdown_text: str) -> str:
        if not markdown_text:
            return ""
            
        text = markdown_text
        text = re.sub(r"```[a-zA-Z0-9]*\n", "", text)
        text = text.replace("```", "")
        text = re.sub(r"!\[([^\]]*)\]\([^)]*\)", r"\1", text)
        text = re.sub(r"\[([^\]]+)\]\([^)]*\)", r"\1", text)
        text = re.sub(r"^\s{0,3}#{1,6}\s*", "", text, flags=re.MULTILINE)
        text = re.sub(r"^\s{0,3}>\s?", "", text, flags=re.MULTILINE)
        text = re.sub(r"^\s*[-*+]\s+", "", text, flags=re.MULTILINE)
        text = re.sub(r"(\*\*|__|\*|_|`)", "", text)
        text = text.replace("|", " ")
        return text.strip()

    def optimize_text(self, text: str) -> str:
        if not text:
            return ""

        text = text.replace("\u00a0", " ").replace("\t", " ")
        text = text.replace("\r\n", "\n").replace("\r", "\n")

        lines = text.split("\n")
        
        line_counts = {}
        for line in lines:
            stripped = line.strip()
            if len(stripped) > 5:
                line_counts[stripped] = line_counts.get(stripped, 0) + 1

        repeated_lines = {l for l, count in line_counts.items() if count > 2}

        cleaned = []
        for line in lines:
            stripped = line.strip()
            
            if stripped in repeated_lines:
                continue

            if (re.fullmatch(r"(?i)page\s+\d+(\s+of\s+\d+)?", stripped) or
                re.fullmatch(r"\d+\s*/\s*\d+", stripped) or
                re.fullmatch(r"-\s*\d+\s*-", stripped) or
                re.fullmatch(r"\[page\s+\d+\]", stripped)):
                continue

            line = re.sub(r"(?i)\bUnnamed:\s*\d+\b", "", line)
            line = re.sub(r"\b(NaN|nan)\b", "", line)

            if re.fullmatch(r"[-=_*~]{3,}", stripped):
                continue

            if re.match(r"^\s*\|?(\s*:?-+:?\s*\|)+\s*$", line):
                continue
            if "|" in line:
                cells = [c.strip() for c in line.split("|")]
                if cells and cells[0] == "":
                    cells = cells[1:]
                if cells and cells[-1] == "":
                    cells = cells[:-1]
                line = " | ".join(cells)

            line = re.sub(r"[ ]{2,}", " ", line).rstrip()
            cleaned.append(line)

        result_lines = []
        blank_run = 0
        for line in cleaned:
            if line.strip() == "":
                blank_run += 1
                if blank_run >= 2:
                    continue
                result_lines.append("")
            else:
                blank_run = 0
                result_lines.append(line)

        return "\n".join(result_lines).strip()

document_service = DocumentService()
